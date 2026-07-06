"""Document generation tools — PDF and Excel with Mitsumi branding.

PDF uses Quicksand font family with proper visual hierarchy,
branded colors, clean tables, and structured layouts.
"""

from __future__ import annotations

import json
import os
import re
from datetime import datetime, timezone
from pathlib import Path

from langchain_core.tools import tool

MITSUMI_LOGO_URL = "https://res.cloudinary.com/dunssu2gi/image/upload/v1767612787/blog-images/tfvwseshobpnx7blnimx.png"
OUTPUT_DIR = Path("generated")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
FONT_DIR = Path(__file__).parent.parent.parent / "assets" / "fonts"

# Brand colours (RGB tuples for reportlab)
BRAND = (79 / 255, 106 / 255, 245 / 255)
DARK = (11 / 255, 15 / 255, 25 / 255)
MUTED = (100 / 255, 116 / 255, 139 / 255)
LIGHT_BG = (241 / 255, 245 / 255, 249 / 255)


def _download_logo() -> str | None:
    logo_path = OUTPUT_DIR / "mitsumi_logo.png"
    if logo_path.exists():
        return str(logo_path)
    try:
        import urllib.request
        urllib.request.urlretrieve(MITSUMI_LOGO_URL, str(logo_path))
        return str(logo_path)
    except Exception:
        return None


def _register_fonts():
    """Register Quicksand font family with reportlab."""
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    fonts = {
        "Quicksand": "Quicksand-Regular.ttf",
        "Quicksand-Bold": "Quicksand-Bold.ttf",
        "Quicksand-Medium": "Quicksand-Medium.ttf",
        "Quicksand-SemiBold": "Quicksand-SemiBold.ttf",
        "Quicksand-Light": "Quicksand-Light.ttf",
    }
    for name, filename in fonts.items():
        path = FONT_DIR / filename
        if path.exists():
            try:
                pdfmetrics.registerFont(TTFont(name, str(path)))
            except Exception:
                pass


@tool
def file_gen(title: str, body: str = "") -> str:
    """Generate a professional Mitsumi-branded PDF report. Pass the title and body content.
    Body can include markdown: ## headings, **bold**, - bullets, 1. numbered lists, | tables |.
    Returns JSON with the file path and download URL."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.colors import HexColor, Color
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        Image, HRFlowable, KeepTogether,
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER

    _register_fonts()

    # Check if Quicksand is available
    heading_font = "Quicksand-Bold"
    body_font = "Quicksand"
    medium_font = "Quicksand-Medium"
    try:
        from reportlab.pdfbase import pdfmetrics
        pdfmetrics.getFont(heading_font)
    except Exception:
        heading_font = "Helvetica-Bold"
        body_font = "Helvetica"
        medium_font = "Helvetica"

    fname = f"{re.sub(r'[^a-z0-9_-]', '_', title.lower())[:60]}.pdf"
    file_path = OUTPUT_DIR / fname

    brand = HexColor("#4F6AF5")
    dark = HexColor("#0B0F19")
    muted_c = HexColor("#64748B")
    light_bg = HexColor("#F1F5F9")
    white = HexColor("#FFFFFF")
    accent = HexColor("#F97316")

    doc = SimpleDocTemplate(
        str(file_path), pagesize=A4,
        rightMargin=22 * mm, leftMargin=22 * mm,
        topMargin=18 * mm, bottomMargin=18 * mm,
    )

    styles = getSampleStyleSheet()

    # Custom styles with Quicksand
    styles.add(ParagraphStyle("MTitle", fontName=heading_font, fontSize=20,
                              textColor=dark, spaceAfter=4, leading=26))
    styles.add(ParagraphStyle("MSub", fontName=body_font, fontSize=9,
                              textColor=muted_c, spaceAfter=14, leading=12))
    styles.add(ParagraphStyle("MH2", fontName=heading_font, fontSize=13,
                              textColor=brand, spaceBefore=18, spaceAfter=8,
                              leading=17, borderPadding=(0, 0, 4, 0)))
    styles.add(ParagraphStyle("MH3", fontName=medium_font, fontSize=11,
                              textColor=dark, spaceBefore=12, spaceAfter=6, leading=15))
    styles.add(ParagraphStyle("MBody", fontName=body_font, fontSize=9.5,
                              textColor=dark, leading=14, spaceAfter=5))
    styles.add(ParagraphStyle("MBullet", fontName=body_font, fontSize=9.5,
                              textColor=dark, leading=14, spaceAfter=3,
                              leftIndent=16, bulletIndent=6))
    styles.add(ParagraphStyle("MFooter", fontName=body_font, fontSize=7.5,
                              textColor=muted_c, alignment=TA_CENTER))

    story: list = []

    # ── Header ──
    logo_path = _download_logo()
    if logo_path and os.path.exists(logo_path):
        try:
            story.append(Image(logo_path, width=34, height=34))
            story.append(Spacer(1, 6))
        except Exception:
            pass

    story.append(Paragraph(title, styles["MTitle"]))
    from app.core.tz import format_datetime
    now = format_datetime()
    story.append(Paragraph(f"Mitsumi Distribution &middot; {now}", styles["MSub"]))
    story.append(HRFlowable(width="100%", color=brand, thickness=2.5, spaceAfter=16))

    # ── Parse body into structured elements ──
    if body:
        lines = body.split("\n")
        i = 0
        while i < len(lines):
            line = lines[i].strip()

            # Empty line
            if not line:
                story.append(Spacer(1, 4))
                i += 1
                continue

            # Heading ##
            h_match = re.match(r"^(#{1,3})\s+(.+)", line)
            if h_match:
                level = len(h_match.group(1))
                text = h_match.group(2).replace("**", "")
                style = "MH2" if level <= 2 else "MH3"
                story.append(Paragraph(text, styles[style]))
                i += 1
                continue

            # Horizontal rule
            if re.match(r"^[-_*]{3,}\s*$", line):
                story.append(HRFlowable(width="100%", color=muted_c, thickness=0.5,
                                        spaceBefore=8, spaceAfter=8))
                i += 1
                continue

            # Table detection
            if "|" in line and i + 1 < len(lines) and re.match(r"^\|?\s*[-:|]+", lines[i + 1].strip()):
                headers = [c.strip().replace("**", "") for c in line.split("|") if c.strip()]
                i += 2  # skip header + separator
                rows = []
                while i < len(lines) and "|" in lines[i]:
                    cells = [c.strip().replace("**", "") for c in lines[i].split("|") if c.strip()]
                    if cells:
                        rows.append(cells)
                    i += 1

                if headers and rows:
                    # Normalize column count
                    max_cols = max(len(headers), max((len(r) for r in rows), default=0))
                    while len(headers) < max_cols:
                        headers.append("")
                    for r in rows:
                        while len(r) < max_cols:
                            r.append("")

                    # Use Paragraph objects so text wraps in cells
                    cell_style = ParagraphStyle("CellText", fontName=body_font, fontSize=8,
                                                leading=10, textColor=dark)
                    header_style = ParagraphStyle("CellHeader", fontName=heading_font, fontSize=8,
                                                  leading=10, textColor=white)

                    table_data = [[Paragraph(h, header_style) for h in headers]]
                    for row in rows:
                        table_data.append([Paragraph(c, cell_style) for c in row])

                    # Proportional widths based on max content length per column
                    col_max = [0] * max_cols
                    for row in [headers] + rows:
                        for ci, cell in enumerate(row):
                            col_max[ci] = max(col_max[ci], len(cell))
                    total = max(sum(col_max), 1)
                    col_widths = [max(doc.width * (cm / total), 40) for cm in col_max]
                    # Scale to fit page
                    scale = doc.width / sum(col_widths)
                    col_widths = [w * scale for w in col_widths]

                    t = Table(table_data, colWidths=col_widths, repeatRows=1)
                    t.setStyle(TableStyle([
                        ("BACKGROUND", (0, 0), (-1, 0), brand),
                        ("TEXTCOLOR", (0, 0), (-1, 0), white),
                        ("FONTNAME", (0, 0), (-1, 0), heading_font),
                        ("FONTSIZE", (0, 0), (-1, 0), 8.5),
                        ("FONTNAME", (0, 1), (-1, -1), body_font),
                        ("FONTSIZE", (0, 1), (-1, -1), 8.5),
                        ("TEXTCOLOR", (0, 1), (-1, -1), dark),
                        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [white, light_bg]),
                        ("GRID", (0, 0), (-1, -1), 0.5, HexColor("#CBD5E1")),
                        ("TOPPADDING", (0, 0), (-1, -1), 6),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                        ("LEFTPADDING", (0, 0), (-1, -1), 8),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                        ("ROUNDEDCORNERS", [4, 4, 4, 4]),
                    ]))
                    story.append(Spacer(1, 6))
                    story.append(t)
                    story.append(Spacer(1, 8))
                continue

            # Unordered list
            if re.match(r"^[-*]\s", line):
                items = []
                while i < len(lines) and re.match(r"^[-*]\s", lines[i].strip()):
                    text = re.sub(r"^[-*]\s+", "", lines[i].strip())
                    text = _clean_md(text)
                    items.append(text)
                    i += 1
                for item in items:
                    story.append(Paragraph(f"&bull;&nbsp;&nbsp;{item}", styles["MBullet"]))
                continue

            # Ordered list
            if re.match(r"^\d+\.\s", line):
                items = []
                while i < len(lines) and re.match(r"^\d+\.\s", lines[i].strip()):
                    text = re.sub(r"^\d+\.\s+", "", lines[i].strip())
                    text = _clean_md(text)
                    items.append(text)
                    i += 1
                for idx, item in enumerate(items, 1):
                    story.append(Paragraph(f"{idx}.&nbsp;&nbsp;{item}", styles["MBullet"]))
                continue

            # Regular paragraph
            clean = _clean_md(line)
            story.append(Paragraph(clean, styles["MBody"]))
            i += 1
    else:
        story.append(Paragraph("Report generated — no additional content provided.", styles["MBody"]))

    # ── Footer ──
    story.append(Spacer(1, 24))
    story.append(HRFlowable(width="100%", color=muted_c, thickness=0.5, spaceAfter=8))
    story.append(Paragraph(
        "Confidential &middot; Mitsumi Distribution &middot; mitsumidistribution.com",
        styles["MFooter"],
    ))

    doc.build(story)
    return json.dumps({"path": str(file_path), "url": f"/static/{fname}", "format": "pdf", "title": title})


def _clean_md(text: str) -> str:
    """Convert markdown bold/italic to reportlab XML tags."""
    text = re.sub(r"\*\*(.+?)\*\*", r"<b>\1</b>", text)
    text = re.sub(r"\*(.+?)\*", r"<i>\1</i>", text)
    text = re.sub(r"`([^`]+)`", r"<font face='Courier' size='8'>\1</font>", text)
    return text


@tool
def excel_export(title: str, data: str = "[]") -> str:
    """Generate a branded Excel spreadsheet. Pass a title and data as a JSON string
    (list of objects or list of lists). If data is empty, an empty spreadsheet is created.
    Returns JSON with the file path."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    fname = f"{re.sub(r'[^a-z0-9_-]', '_', title.lower())[:60]}.xlsx"
    file_path = OUTPUT_DIR / fname

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4F6AF5", end_color="4F6AF5", fill_type="solid")
    body_font = Font(name="Calibri", size=10, color="0B0F19")
    alt_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    try:
        parsed = json.loads(data) if isinstance(data, str) else data
    except Exception:
        parsed = [{"error": "Could not parse data", "raw": data[:500]}]

    if isinstance(parsed, list) and parsed and isinstance(parsed[0], dict):
        headers = list(parsed[0].keys())
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 1))
        title_cell = ws.cell(row=1, column=1, value=f"Mitsumi Distribution — {title}")
        title_cell.font = Font(name="Calibri", bold=True, size=14, color="4F6AF5")
        from app.core.tz import format_short
        ws.cell(row=2, column=1, value=f"Generated {format_short()}")
        ws.cell(row=2, column=1).font = Font(name="Calibri", size=9, color="64748B")

        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=h.replace("_", " ").title())
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left")
            cell.border = thin_border
            ws.column_dimensions[cell.column_letter].width = max(len(h) * 1.5, 12)

        for row_idx, row_data in enumerate(parsed, 5):
            for col_idx, h in enumerate(headers, 1):
                val = row_data.get(h, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=str(val) if val is not None else "")
                cell.font = body_font
                cell.border = thin_border
                if (row_idx - 5) % 2 == 1:
                    cell.fill = alt_fill
    else:
        ws.cell(row=1, column=1, value=title)
        ws.cell(row=2, column=1, value=str(parsed)[:32000])

    wb.save(str(file_path))
    return json.dumps({"path": str(file_path), "url": f"/static/{fname}", "format": "xlsx", "title": title})
